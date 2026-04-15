#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simulation complète du pipeline d'import fournisseurs.

1. Génère 3 faux fichiers xlsx (un par fournisseur, formats différents)
2. Passe chaque fichier dans son adaptateur
3. Matche les EAN sur un catalogue produit fictif
4. Simule le résultat qui serait écrit dans seller_ids Odoo
"""

import os
import openpyxl
from importers import IMPORTERS

OUT_DIR = os.path.join(os.path.dirname(__file__), "sample_data")
os.makedirs(OUT_DIR, exist_ok=True)


# ---------------------------------------------------------------------------
# Catalogue produit fictif  (EAN → nom produit)
# Remplace par une vraie requête Odoo en production
# ---------------------------------------------------------------------------

PRODUCT_CATALOG = {
    "3700123456789": "Clavier mécanique USB",
    "3700123456796": "Souris sans fil",
    "3700123456802": "Écran 27\" Full HD",
    "3700123456819": "Webcam HD 1080p",
    "3700123456826": "Casque audio Bluetooth",
    "3700123456833": "Hub USB-C 7 ports",
}


# ---------------------------------------------------------------------------
# Générateurs de fichiers xlsx de test
# ---------------------------------------------------------------------------

def make_alpha_xlsx(path):
    """Format Alpha : onglet 1, headers ligne 1, prix float."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tarifs"
    ws.append(["Code EAN", "Désignation", "Prix HT", "Promotion", "Qté mini"])
    ws.append(["3700123456789", "Clavier mec.",  38.50, "Non", 1])
    ws.append(["3700123456796", "Souris SF",     18.00, "Oui", 1])   # promo
    ws.append(["3700123456802", "Ecran 27",     145.00, "Non", 1])
    ws.append(["9999999999999", "Produit inconnu", 5.00, "Non", 1])  # EAN absent du catalogue
    ws.append([None,            "Ligne sans EAN",  3.00, "Non", 1])  # ignorée
    wb.save(path)


def make_beta_xlsx(path):
    """Format Beta : données sur onglet 2, headers ligne 3, prix "xx,xx EUR"."""
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "Infos"               # onglet 0 — inutilisé
    ws = wb.create_sheet("Catalogue") # onglet 1 — lu par l'adaptateur
    # 2 lignes vides avant les headers (header_row=2)
    ws.append(["Export catalogue fournisseur Beta"])
    ws.append(["Généré le 2026-04-15"])
    ws.append(["GTIN", "Libellé", "Tarif", "Offre spéciale"])
    ws.append(["3700123456819", "Webcam HD",    "22,90 EUR", "Non"])
    ws.append(["3700123456826", "Casque BT",    "55,00 EUR", "Oui"])  # promo
    ws.append(["3700123456833", "Hub USB-C",    "19,50 EUR", "Non"])
    ws.append(["3700123456789", "Clavier mec.", "41,00 EUR", "Non"])  # aussi chez Alpha (plus cher)
    wb.save(path)


def make_gamma_xlsx(path):
    """Format Gamma : headers lowercase, EAN sans zéros initiaux parfois."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "price_list"
    ws.append(["ean", "product_name", "unit_price"])
    ws.append(["3700123456796", "Souris SF",     21.00])   # aussi chez Alpha (plus cher, non promo)
    ws.append(["3700123456802", "Ecran 27",      139.90])  # moins cher qu'Alpha
    ws.append(["370012345678",  "EAN court",      10.00])  # sera zero-paddé → 0370012345678
    wb.save(path)


# ---------------------------------------------------------------------------
# Matching EAN → produit Odoo (mock)
# ---------------------------------------------------------------------------

def match_products(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Sépare les lignes matchées des lignes sans produit correspondant."""
    matched, unmatched = [], []
    for row in rows:
        if row["ean"] in PRODUCT_CATALOG:
            matched.append({**row, "product_name": PRODUCT_CATALOG[row["ean"]]})
        else:
            unmatched.append(row)
    return matched, unmatched


# ---------------------------------------------------------------------------
# Affichage
# ---------------------------------------------------------------------------

def print_results(matched: list[dict], unmatched: list[dict], supplier: str):
    SEP = "-" * 62
    print(f"\n{'=' * 62}")
    print(f"  {supplier}")
    print(SEP)
    print(f"  {'EAN':<15} {'Produit':<25} {'Prix':>8}  {'Promo'}")
    print(SEP)
    for r in matched:
        promo = "OUI" if r["is_promo"] else ""
        print(f"  {r['ean']:<15} {r['product_name']:<25} {r['price']:>7.2f}€  {promo}")
    if unmatched:
        print(f"\n  ⚠  {len(unmatched)} EAN non trouvé(s) dans le catalogue Odoo :")
        for r in unmatched:
            print(f"     • {r['ean']}  ({r['price']:.2f}€)")
    print(f"{'=' * 62}")


def print_final_result(all_matched: list[dict]):
    """Simule ce qui serait écrit dans seller_ids : 1 ligne par produit+fournisseur."""
    print(f"\n{'#' * 62}")
    print("  RÉSULTAT FINAL — seller_ids qui seraient créés/mis à jour")
    print(f"{'#' * 62}")

    # Regrouper par EAN
    by_ean = {}
    for r in all_matched:
        by_ean.setdefault(r["ean"], []).append(r)

    for ean, rows in sorted(by_ean.items()):
        product_name = rows[0]["product_name"]
        print(f"\n  {product_name}  [{ean}]")
        non_promo = [r for r in rows if not r["is_promo"]]
        for r in sorted(rows, key=lambda x: x["price"]):
            marker = " ← RETENU" if non_promo and r == min(non_promo, key=lambda x: x["price"]) else ""
            promo  = " [PROMO]" if r["is_promo"] else ""
            print(f"    {r['supplier_name']:<22} {r['price']:>7.2f}€{promo}{marker}")


# ---------------------------------------------------------------------------
# Pipeline principal
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    files = {
        "alpha": os.path.join(OUT_DIR, "fournisseur_alpha.xlsx"),
        "beta":  os.path.join(OUT_DIR, "fournisseur_beta.xlsx"),
        "gamma": os.path.join(OUT_DIR, "fournisseur_gamma.xlsx"),
    }

    print("Génération des fichiers de test...")
    make_alpha_xlsx(files["alpha"])
    make_beta_xlsx(files["beta"])
    make_gamma_xlsx(files["gamma"])
    print("OK\n")

    all_matched = []

    for key, filepath in files.items():
        importer = IMPORTERS[key]
        print(f"Import : {filepath}")
        rows = importer.load(filepath)
        matched, unmatched = match_products(rows)
        print_results(matched, unmatched, importer.supplier_name)
        all_matched.extend(matched)

    print_final_result(all_matched)
